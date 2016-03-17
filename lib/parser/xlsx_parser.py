from datetime import datetime

from dateutil import parser as date_parser

from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill

from lib.models.roster import Roster
from lib.models.shift import Shift
from lib.parser import BaseParser
from lib.parser.helper.cell import Cell
from lib.config import PEOPLE_START_CELL, PEOPLE_END_CELL, SHIFT_ROW_LIMIT, TIMES_START_CELL, TIMES_END_CELL


class XLSXParser(BaseParser):
    def parse(self):
        workbook = load_workbook(filename=self.file_name)
        ws = workbook.active

        times = self.extract_times(ws)
        people = self.extract_people(ws)
        shifts = self.extract_shifts(ws, times, people)

        return Roster(shifts=shifts, people=people)

    @staticmethod
    def extract_times(ws):
        times = {}
        cells = Cell.row_range(start_cell=TIMES_START_CELL, end_cell=TIMES_END_CELL)

        try:
            current_date = None
            for cell in cells:
                _, row = Cell.split_col_row(cell)
                date_value = ws[str(cell)].value

                new_date = validate_date(date_value)
                current_date = new_date if new_date is not None else current_date

                time_value = ws[str(cell.right())].value

                if time_value:
                    start_time, end_time = validate_times(time_value)

                    try:
                        start_datetime = combine_date_and_time(current_date, start_time)
                        end_datetime = combine_date_and_time(current_date, end_time)
                    except Exception as e:
                        print('WARNING: Time Cell: {}. Error: {}'.format(cell.right(), e))

                    times[row] = Shift(start_time=start_datetime, end_time=end_datetime)
        except Exception as e:
            print('Error extracting times. Cell={},Error={}'.format(cell, e))
            raise Exception(e)

        return times

    @staticmethod
    def extract_people(ws):
        people = []
        cells = Cell.col_range(start_cell=PEOPLE_START_CELL, end_cell=PEOPLE_END_CELL)
        for cell in cells:
            _, row = Cell.split_col_row(cell)

            name = ws[str(cell)].value
            if name:
                people.append(name)
            else:
                break

        return people

    @staticmethod
    def extract_shifts(ws, times, people):
        shifts = []

        for p in range(len(people)):
            start_cell = Cell.from_cell(row=PEOPLE_START_CELL.down().row, col=PEOPLE_START_CELL.right(p).col)
            end_cell = Cell.from_cell(row=SHIFT_ROW_LIMIT, col=start_cell.col)

            cells = Cell.row_range(start_cell=start_cell, end_cell=end_cell)
            while True:
                cells, shift = XLSXParser.extract_shift(
                    ws=ws,
                    times=times,
                    person=people[p],
                    cells=cells)

                if shift is not None:
                    shifts.append(shift)
                elif len(cells) == 0:
                    break

        return shifts

    @staticmethod
    def extract_shift(ws, times, person, cells):
        start_time = end_time = location = None

        i = 0
        for cell in cells:
            i = i + 1
            value = ws[str(cell)].value
            fill = ws[str(cell)].fill

            location = XLSXParser.location_from_cell(ws, cell)

            if location:
                start_time = times[cell.row].start_time if start_time is None else start_time
                end_time = times[cell.row].end_time

            next_location = XLSXParser.location_from_cell(ws, cell.down())
            if not next_location or location != next_location:
                break

        if None not in [start_time, end_time, location]:
            shift = Shift(
                start_time=start_time,
                end_time=end_time,
                person=person,
                location=location
            )
        else:
            shift = None

        return cells[i+1:], shift

    @staticmethod
    def location_from_cell(ws, cell):
        fill = ws[str(cell)].fill

        if len(str(fill)) == 0 or fill.fgColor.type != 'indexed':
            return None

        return XLSXParser.colour_to_location(fill.fgColor.index)

    @staticmethod
    def apply_filter(text):
        split = text.split(' ', maxsplit=2)
        command = split[0]
        args = split[1]

        if command in ['start']:
            cmd_start(args)
        elif command in ['end', 'finish']:
            cmd_end(args)
        elif command in ['wk', 'wks']:
            cmd_wks(args)
        else:
            cmd_range(text)

        def cmd_start(args):
            pass

        def cmd_end(args):
            pass

        def cmd_range(args):
            pass

        def cmd_wks(args):
            pass


def validate_date(date_str):
    try:
        return date_parser.parse(date_str)
    except Exception as e:
        return None


def validate_times(time_str):
    try:
        time_split = time_str.split('-', maxsplit=2)
        times = list(map(date_parser.parse, time_split))

        return times[0], times[1]
    except Exception as e:
        print('WARNING: Could not validate time string [{}]'.format(time_str))
        return None, None


def combine_date_and_time(date, time):
    try:
        return datetime(year=date.year, month=date.month, day=date.day,
                        hour=time.hour, minute=time.minute)
    except Exception as e:
        raise Exception('Failed to combine date and time. Date: [{}] Time: [{}]'.format(date, time))

